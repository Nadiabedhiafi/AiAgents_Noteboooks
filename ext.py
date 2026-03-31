# “””
Credit Memo Analyzer

Extracts structured sections (A1, A2, B1, etc.) from multiple credit memo PDFs,
groups identical sections across files, then uses Claude to:

- Identify common points across sections of the same type
- Generate general questions that the section answers
- Produce a description of what each section type covers

Dependencies:
pip install pdfplumber anthropic openpyxl

Usage:
python credit_memo_analyzer.py –pdf_dir ./pdfs –output results.xlsx
python credit_memo_analyzer.py –pdf_dir ./pdfs –output results.xlsx –api_key sk-ant-…
“””

import os
import re
import json
import argparse
import pdfplumber
import anthropic
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────

# 1. SECTION EXTRACTION FROM PDFs

# ─────────────────────────────────────────────

# Pattern matching: (A1), (A2), (B1)… letter + at least one digit — skips lone-letter sections like (A) or (B)

SECTION_PATTERN = re.compile(r’^(([A-Z]\d+))\s*[-–—:•]?\s*(.+?)\s*$’, re.IGNORECASE | re.MULTILINE)

def extract_text_from_pdf(pdf_path: str) -> str:
“”“Extract full text from a PDF, page by page.”””
text_parts = []
with pdfplumber.open(pdf_path) as pdf:
for page in pdf.pages:
page_text = page.extract_text()
if page_text:
text_parts.append(page_text)
return “\n”.join(text_parts)

def parse_sections(text: str) -> dict[str, dict]:
“””
Walk through the text line by line and split into sections.
Returns a dict: { “A1”: {“title”: “…”, “content”: “…”}, … }
“””
sections = {}
current_key = None
current_title = “”
current_lines = []

```
for line in text.splitlines():
    line_stripped = line.strip()
    match = SECTION_PATTERN.match(line_stripped)

    if match:
        # Save the previous section before starting a new one
        if current_key:
            sections[current_key] = {
                "title": current_title,
                "content": "\n".join(current_lines).strip()
            }
        # Start a new section
        current_key = match.group(1).upper()
        current_title = match.group(2).strip()
        current_lines = []
    elif current_key:
        current_lines.append(line_stripped)

# Save the last section
if current_key:
    sections[current_key] = {
        "title": current_title,
        "content": "\n".join(current_lines).strip()
    }

return sections
```

def extract_sections_from_folder(pdf_dir: str) -> dict[str, list[dict]]:
“””
Scan all PDFs in a folder and group sections by type.
Returns:
{
“A1”: [ {“source”: “file.pdf”, “title”: “…”, “content”: “…”}, … ],
“B2”: [ … ],
…
}
“””
grouped: dict[str, list[dict]] = defaultdict(list)
pdf_files = list(Path(pdf_dir).glob(”*.pdf”))

```
if not pdf_files:
    raise FileNotFoundError(f"No PDF files found in: {pdf_dir}")

print(f"📂 {len(pdf_files)} PDF(s) found in '{pdf_dir}'")

for pdf_path in sorted(pdf_files):
    print(f"  📄 Reading: {pdf_path.name}")
    try:
        text = extract_text_from_pdf(str(pdf_path))
        sections = parse_sections(text)
        if not sections:
            print(f"     ⚠️  No sections detected in {pdf_path.name}")
        for key, data in sections.items():
            grouped[key].append({
                "source": pdf_path.name,
                "title": data["title"],
                "content": data["content"]
            })
            print(f"     ✅ Section {key}: '{data['title'][:60]}'")
    except Exception as e:
        print(f"     ❌ Error processing {pdf_path.name}: {e}")

return dict(grouped)
```

# ─────────────────────────────────────────────

# 2. CLAUDE ANALYSIS

# ─────────────────────────────────────────────

def _call_claude(client: anthropic.Anthropic, prompt: str) -> str:
“”“Send a single prompt to Claude and return the raw text response.”””
response = client.messages.create(
model=“claude-opus-4-5”,
max_tokens=1500,
messages=[{“role”: “user”, “content”: prompt}]
)
raw = response.content[0].text.strip()
# Strip any accidental `json ... ` fences
raw = re.sub(r’^`(?:json)?\s*', '', raw) raw = re.sub(r'\s*`$’, ‘’, raw)
return raw

def summarize_single_entry(client: anthropic.Anthropic, section_key: str, entry: dict) -> str:
“””
Call 1 — One LLM call per individual PDF entry.
Extracts a compact bullet-point summary of that single section,
keeping the token footprint small before the synthesis step.
“””
prompt = f””“You are reading a ({section_key}) section from a credit memo for company “{entry[‘source’]}”.

Title: {entry[‘title’]}
Content:
{entry[‘content’]}

Analyze this section and return a structured summary with the following bullet points:

CONTENT SUMMARY (3-5 bullets):

- Key information and data points found in this section

CONTENT TYPE:

- Identify what type of content this section contains. For each type present, specify what it represents:
  - Descriptive text: briefly state the topic it describes
  - Table: state what data the table presents (e.g., financial figures, comparisons)
  - Chart / Graph: state what the chart visualizes (e.g., revenue trend, risk distribution)
  - Image / Diagram: state what the image or diagram illustrates

Return ONLY the structured bullet points above, no preamble or conclusion.
“””
return _call_claude(client, prompt)

def synthesize_section(client: anthropic.Anthropic, section_key: str,
summaries: list[dict]) -> dict:
“””
Call 2 — One final LLM call that receives only the compact summaries
(not the full raw content) and produces the consolidated JSON output.
This avoids hitting token limits when there are many PDFs.
“””
blocks = “\n\n”.join(
f”— Source {i+1}: {s[‘source’]} —\n{s[‘summary’]}”
for i, s in enumerate(summaries)
)

```
prompt = f"""You are synthesizing ({section_key}) section summaries extracted from {len(summaries)} different credit memos.
```

Here are the individual summaries:

{blocks}

Respond with valid JSON ONLY — no text or Markdown fences — using exactly these fields:

{{
“section_key”: “{section_key}”,
“description”: “A clear and concise description of what this section type covers (2-4 sentences).”,
“content_types”: {{
“descriptive_text”: “What the descriptive text in this section typically covers, or null if absent.”,
“table”: “What the table(s) in this section typically present, or null if absent.”,
“chart”: “What the chart(s) or graph(s) in this section typically visualize, or null if absent.”,
“image”: “What the image(s) or diagram(s) in this section typically illustrate, or null if absent.”
}},
“common_points”: [
“Common point 1 shared across all these sections”,
“Common point 2”,
“…”
],
“suggested_questions”: [
“General question 1 that this section answers”,
“General question 2”,
“General question 3”,
“…”
]
}}
“””
raw = _call_claude(client, prompt)
try:
return json.loads(raw)
except json.JSONDecodeError:
# Fallback if the JSON response is malformed
return {
“section_key”: section_key,
“description”: “Failed to parse Claude’s response.”,
“common_points”: [],
“suggested_questions”: [],
“raw_response”: raw
}

def analyze_section_with_claude(
client: anthropic.Anthropic,
section_key: str,
entries: list[dict]
) -> dict:
“””
Two-step analysis to stay within token limits:
Step A — one small LLM call per PDF entry to summarize it individually.
Step B — one final LLM call on the compact summaries to synthesize the result.
“””
# Step A: summarize each entry independently
summaries = []
for i, entry in enumerate(entries, 1):
print(f”       📝 Summarizing entry {i}/{len(entries)}: {entry[‘source’]}…”)
summary_text = summarize_single_entry(client, section_key, entry)
summaries.append({“source”: entry[“source”], “summary”: summary_text})

```
# Step B: synthesize all summaries into the final JSON
print(f"       🔗 Synthesizing {len(summaries)} summaries...")
return synthesize_section(client, section_key, summaries)
```

def analyze_all_sections(
grouped_sections: dict[str, list[dict]],
api_key: str | None = None
) -> list[dict]:
“”“Analyze all section types and return the list of results.”””
client = anthropic.Anthropic(api_key=api_key) if api_key else anthropic.Anthropic()

```
results = []
sorted_keys = sorted(grouped_sections.keys())

print(f"\n🤖 Analyzing {len(sorted_keys)} section type(s) with Claude...")

for key in sorted_keys:
    entries = grouped_sections[key]
    print(f"  🔍 Section ({key}) — {len(entries)} occurrence(s)...")

    if not entries:
        continue

    analysis = analyze_section_with_claude(client, key, entries)
    analysis["occurrences"] = len(entries)
    analysis["sources"] = [e["source"] for e in entries]
    # Collect distinct titles found across PDFs for this section key
    analysis["title"] = next((e["title"] for e in entries if e.get("title")), "")
    results.append(analysis)

    print(f"     ✅ Description: {analysis.get('description', '')[:80]}…")

return results
```

# ─────────────────────────────────────────────

# 3. EXCEL EXPORT

# ─────────────────────────────────────────────

# ── Color palette ─────────────────────────────

COLOR_HEADER_BG  = “1F3864”   # Dark navy (header background)
COLOR_HEADER_FG  = “FFFFFF”   # White (header text)
COLOR_SECTION_BG = “2E75B6”   # Medium blue (section title banner)
COLOR_SECTION_FG = “FFFFFF”
COLOR_LABEL_BG   = “D6E4F0”   # Light blue (label cells)
COLOR_ALT_ROW    = “EBF3FB”   # Very light blue (alternating rows)
COLOR_BORDER     = “A0C4E8”

THIN_BORDER = Border(
left=Side(style=“thin”, color=COLOR_BORDER),
right=Side(style=“thin”, color=COLOR_BORDER),
top=Side(style=“thin”, color=COLOR_BORDER),
bottom=Side(style=“thin”, color=COLOR_BORDER),
)

def _cell(ws, row, col, value=””, bold=False, fg=None, bg=None,
wrap=True, align_v=“top”, align_h=“left”, size=11, border=True, italic=False):
“”“Helper to write and style a single cell.”””
c = ws.cell(row=row, column=col, value=value)
c.font = Font(name=“Arial”, bold=bold, italic=italic,
color=fg or “000000”, size=size)
if bg:
c.fill = PatternFill(“solid”, start_color=bg)
c.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
if border:
c.border = THIN_BORDER
return c

def _merge_row(ws, row, col_start, col_end, value=””, bold=False,
fg=None, bg=None, size=11):
“”“Merge columns on a row and apply styling.”””
ws.merge_cells(
start_row=row, start_column=col_start,
end_row=row, end_column=col_end
)
c = ws.cell(row=row, column=col_start, value=value)
c.font = Font(name=“Arial”, bold=bold, color=fg or “000000”, size=size)
if bg:
c.fill = PatternFill(“solid”, start_color=bg)
c.alignment = Alignment(horizontal=“left”, vertical=“center”, wrap_text=True)
c.border = THIN_BORDER
return c

def build_summary_sheet(wb: Workbook, results: list[dict]):
“””
Sheet 1 — Overview: one row per section type with
section | occurrences | sources | short description | counts.
“””
ws = wb.active
ws.title = “Overview”

```
# Column headers and widths
headers = ["Section", "Title", "Occurrences", "PDF Sources", "Description", "Common Points", "Questions"]
widths   = [10,        35,       14,          40,            60,            16,              14]

for col, (h, w) in enumerate(zip(headers, widths), 1):
    _cell(ws, 1, col, h, bold=True, fg=COLOR_HEADER_FG,
          bg=COLOR_HEADER_BG, align_h="center", size=11)
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 22

for i, r in enumerate(results, 2):
    bg = COLOR_ALT_ROW if i % 2 == 0 else None
    _cell(ws, i, 1, f"({r.get('section_key', '')})", bold=True, bg=bg, align_h="center")
    _cell(ws, i, 2, r.get("title", ""), bg=bg)
    _cell(ws, i, 3, r.get("occurrences", 0), bg=bg, align_h="center")
    _cell(ws, i, 4, ", ".join(r.get("sources", [])), bg=bg)
    _cell(ws, i, 5, r.get("description", ""), bg=bg)
    _cell(ws, i, 6, len(r.get("common_points", [])), bg=bg, align_h="center")
    _cell(ws, i, 7, len(r.get("suggested_questions", [])), bg=bg, align_h="center")
    ws.row_dimensions[i].height = 40

ws.freeze_panes = "A2"
```

def build_detail_sheet(wb: Workbook, results: list[dict]):
“””
Sheet 2 — Full detail: for each section type, a block with
description, numbered common points, and numbered suggested questions.
“””
ws = wb.create_sheet(“Section Details”)
ws.column_dimensions[“A”].width = 22
ws.column_dimensions[“B”].width = 100

```
row = 1

# Main title row
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
c = ws.cell(row=row, column=1, value="Detailed Section Analysis — Credit Memos")
c.font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[row].height = 30
row += 2

for r in results:
    key  = r.get("section_key", "?")
    occ  = r.get("occurrences", 0)
    srcs = ", ".join(r.get("sources", []))
    desc = r.get("description", "")
    pts  = r.get("common_points", [])
    qs   = r.get("suggested_questions", [])

    # Section banner row
    _merge_row(ws, row, 1, 2,
               value=f"  Section ({key})  |  {r.get('title', '')}   —   {occ} occurrence(s)",
               bold=True, fg=COLOR_SECTION_FG, bg=COLOR_SECTION_BG, size=12)
    ws.row_dimensions[row].height = 24
    row += 1

    # Sources row
    _cell(ws, row, 1, "Sources", bold=True, bg=COLOR_LABEL_BG, align_h="right")
    _cell(ws, row, 2, srcs)
    ws.row_dimensions[row].height = 18
    row += 1

    # Description row
    _cell(ws, row, 1, "Description", bold=True, bg=COLOR_LABEL_BG, align_h="right")
    _cell(ws, row, 2, desc)
    ws.row_dimensions[row].height = max(18, min(80, 18 * (len(desc) // 90 + 1)))
    row += 1

    # Content types rows — one sub-row per type that is present
    ct = r.get("content_types", {})
    ct_labels = {
        "descriptive_text": "📝 Descriptive Text",
        "table":            "📊 Table",
        "chart":            "📈 Chart / Graph",
        "image":            "🖼️  Image / Diagram",
    }
    first = True
    for ct_key, ct_label in ct_labels.items():
        ct_value = ct.get(ct_key)
        if ct_value and ct_value.lower() not in ("null", "none", "absent", ""):
            label_text = "Content Types" if first else ""
            _cell(ws, row, 1, label_text, bold=True, bg=COLOR_LABEL_BG, align_h="right")
            _cell(ws, row, 2, f"{ct_label}: {ct_value}")
            ws.row_dimensions[row].height = 20
            row += 1
            first = False

    # Common points rows
    _cell(ws, row, 1, "Common Points", bold=True, bg=COLOR_LABEL_BG, align_h="right")
    if pts:
        _cell(ws, row, 2, f"1. {pts[0]}")
        ws.row_dimensions[row].height = 18
        row += 1
        for j, pt in enumerate(pts[1:], 2):
            _cell(ws, row, 1, "", bg=COLOR_LABEL_BG)
            _cell(ws, row, 2, f"{j}. {pt}")
            ws.row_dimensions[row].height = 18
            row += 1
    else:
        _cell(ws, row, 2, "(none)")
        row += 1

    # Suggested questions rows
    _cell(ws, row, 1, "Suggested Questions", bold=True, bg=COLOR_LABEL_BG, align_h="right")
    if qs:
        _cell(ws, row, 2, f"1. {qs[0]}")
        ws.row_dimensions[row].height = 18
        row += 1
        for j, q in enumerate(qs[1:], 2):
            _cell(ws, row, 1, "", bg=COLOR_LABEL_BG)
            _cell(ws, row, 2, f"{j}. {q}")
            ws.row_dimensions[row].height = 18
            row += 1
    else:
        _cell(ws, row, 2, "(none)")
        row += 1

    row += 1  # Blank spacer row between sections
```

def build_questions_sheet(wb: Workbook, results: list[dict]):
“””
Sheet 3 — Questions index: all questions listed with their associated section.
Useful as a quick-reference lookup table.
“””
ws = wb.create_sheet(“Questions Index”)
ws.column_dimensions[“A”].width = 12
ws.column_dimensions[“B”].width = 90
ws.column_dimensions[“C”].width = 55

```
headers = ["Section", "Question", "Section Description"]
for col, h in enumerate(headers, 1):
    _cell(ws, 1, col, h, bold=True, fg=COLOR_HEADER_FG,
          bg=COLOR_HEADER_BG, align_h="center")
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

row = 2
for r in results:
    key  = r.get("section_key", "?")
    desc = r.get("description", "")
    qs   = r.get("suggested_questions", [])
    for i, q in enumerate(qs):
        bg = COLOR_ALT_ROW if row % 2 == 0 else None
        _cell(ws, row, 1, f"({key})", bold=True, bg=bg, align_h="center")
        _cell(ws, row, 2, q, bg=bg)
        # Show description only on the first question row for this section
        _cell(ws, row, 3, desc if i == 0 else "", bg=bg, italic=(i > 0))
        ws.row_dimensions[row].height = 22
        row += 1
```

def save_to_excel(results: list[dict], output_path: str):
“”“Create the Excel file with all 3 sheets and save it.”””
wb = Workbook()
build_summary_sheet(wb, results)
build_detail_sheet(wb, results)
build_questions_sheet(wb, results)
wb.save(output_path)
print(f”\n💾 Excel file saved: {output_path}”)

def print_summary(results: list[dict]):
“”“Print a human-readable summary to the terminal.”””
print(”\n” + “=” * 60)
print(”  CREDIT MEMO ANALYSIS — SUMMARY”)
print(”=” * 60)
for r in results:
key  = r.get(“section_key”, “?”)
occ  = r.get(“occurrences”, 0)
desc = r.get(“description”, “”)
pts  = r.get(“common_points”, [])
qs   = r.get(“suggested_questions”, [])
print(f”\n> Section ({key})  —  {occ} occurrence(s)”)
print(f”  {desc}”)
if pts:
print(f”  {len(pts)} common point(s)”)
if qs:
print(f”  {len(qs)} suggested question(s)”)
print(”\n” + “=” * 60)

# ─────────────────────────────────────────────

# 4. ENTRY POINT

# ─────────────────────────────────────────────

def main():
parser = argparse.ArgumentParser(
description=“Analyze structured sections from credit memo PDFs using Claude.”
)
parser.add_argument(
“–pdf_dir”,
required=True,
help=“Folder containing the PDF files (e.g. ./pdfs)”
)
parser.add_argument(
“–output”,
default=“credit_memo_analysis.xlsx”,
help=“Output Excel file path (default: credit_memo_analysis.xlsx)”
)
parser.add_argument(
“–api_key”,
default=None,
help=“Anthropic API key (optional if ANTHROPIC_API_KEY env var is set)”
)

```
args = parser.parse_args()

# Step 1: Extract sections from all PDFs
grouped_sections = extract_sections_from_folder(args.pdf_dir)

if not grouped_sections:
    print("No sections extracted. Check the format of your PDFs.")
    return

print(f"\nSections detected: {sorted(grouped_sections.keys())}")

# Step 2: Analyze each section type with Claude
results = analyze_all_sections(grouped_sections, api_key=args.api_key)

# Step 3: Export to Excel and print terminal summary
save_to_excel(results, args.output)
print_summary(results)
```

if **name** == “**main**”:
main()